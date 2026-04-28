from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import settings


HEADER_MAP = {
    "Opportunity_ID": "opportunity_id",
    "Client": "client",
    "CIT": "cit",
    "Opportunity_Name": "opportunity_name",
    "BFF_Status": "bff_status",
    "Stage": "stage",
    "Generating_Partner": "generating_partner",
    "Hany_Involvement_Type": "hany_involvement_type",
    "Hany_Is_Core_Pursuit": "hany_is_core_pursuit",
    "Fees_Value": "fees_value",
    "Fees_Currency": "fees_currency",
    "Fees_Value_USD": "fees_value_usd",
    "Probability": "probability",
    "Industry_Sector": "industry_sector",
    "Capability_Teams": "capability_teams",
    "Opportunity_Source": "opportunity_source",
    "Country": "country",
    "City_or_State": "city_or_state",
    "Estimated_Start_Date": "estimated_start_date",
    "Duration_Weeks": "duration_weeks",
    "Software_or_AI_Involved": "software_or_ai_involved",
    "Created_On": "created_on",
}


def _normalize(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def load_opportunities() -> list[dict[str, Any]]:
    workbook_path = Path(settings.opportunities_workbook_path)
    if not workbook_path.exists():
        return []

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Opportunities"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    mapped_headers = [HEADER_MAP.get(header, str(header).lower()) for header in headers]
    opportunities: list[dict[str, Any]] = []
    for row in rows:
        if not any(row):
            continue
        record = {
            key: _normalize(value)
            for key, value in zip(mapped_headers, row)
        }
        record["opportunity_id"] = str(record.get("opportunity_id") or "").strip()
        record["client"] = str(record.get("client") or "").strip()
        record["opportunity_name"] = str(record.get("opportunity_name") or "").strip()
        if not record["opportunity_id"] or not record["client"] or not record["opportunity_name"]:
            continue
        opportunities.append(record)
    return opportunities


def build_opportunity_summary(opportunities: list[dict[str, Any]]) -> dict[str, Any]:
    by_stage: dict[str, int] = {}
    total_fees_usd = 0.0
    weighted_fees_usd = 0.0
    core_pursuits = 0

    for item in opportunities:
        stage = item.get("stage") or "Unknown"
        by_stage[stage] = by_stage.get(stage, 0) + 1
        fees = float(item.get("fees_value_usd") or 0)
        probability = float(item.get("probability") or 0)
        total_fees_usd += fees
        weighted_fees_usd += fees * probability
        if item.get("hany_is_core_pursuit"):
            core_pursuits += 1

    return {
        "count": len(opportunities),
        "core_pursuits": core_pursuits,
        "total_fees_usd": total_fees_usd,
        "weighted_fees_usd": weighted_fees_usd,
        "by_stage": by_stage,
    }
